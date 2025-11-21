from flask import Blueprint, request, jsonify, send_from_directory, send_file
from datetime import datetime, date
import jwt
import os
from dotenv import load_dotenv, dotenv_values
from supabase import create_client, Client
from pathlib import Path
from io import BytesIO
import zipfile
import json
import requests
import pandas as pd
from typing import Literal
from flask import current_app

# 환경 변수 로드
load_dotenv()

# 환경변수 안전 로더(BOM/공백 대응)
def _get_env_safe(key: str, default: str = "") -> str:
    try:
        val = os.getenv(key)
        if isinstance(val, str) and val.strip() != "":
            return val.strip()
        # .env 직접 파싱하여 BOM 혹은 비정상 키명 보정
        cfg = {}
        try:
            cfg = dotenv_values('.env') or {}
        except Exception:
            cfg = {}
        # 1) 정확한 키
        if key in cfg and str(cfg[key] or '').strip() != '':
            return str(cfg[key]).strip()
        # 2) BOM이 앞에 붙은 키(\ufeff)
        bom_key = "\ufeff" + key
        if bom_key in cfg and str(cfg[bom_key] or '').strip() != '':
            return str(cfg[bom_key]).strip()
        # 3) 선행 비문자 제거 후 일치하는 키(예외적 상황 방어)
        for k, v in cfg.items():
            ks = str(k or '').strip()
            if ks.endswith(key) and str(v or '').strip() != '':
                return str(v).strip()
        return default
    except Exception:
        return default

# auth.py와 동일한 기본 비밀키 정책 적용 (토큰 검증 시 일관성)
SECRET_KEY = os.getenv('FLASK_SECRET_KEY', 'dev-secret-key-change-in-production')
EMERGENCY_ADMIN_CODE = _get_env_safe('EMERGENCY_ADMIN_CODE', '')

# Blueprint는 모든 라우트 정의보다 먼저 선언되어야 합니다.
sites_bp = Blueprint('sites', __name__)

# Supabase 클라이언트 초기화
supabase_url = os.getenv('SUPABASE_URL')
supabase_key = os.getenv('SUPABASE_ANON_KEY')
supabase_service_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')  # Storage 전용 사용 권장

try:
    print(f"[INFO] Supabase URL: {supabase_url}")
    print(f"[INFO] Supabase Key: {supabase_key[:20]}..." if supabase_key else "[WARN] Supabase Key 없음")
except Exception:
    pass

if not supabase_url or not supabase_key:
    try:
        print("[WARN] Supabase 환경 변수가 설정되지 않았습니다! 더미 데이터로 실행됩니다!.")
    except Exception:
        pass
    
    # 더미 Supabase 클라이언트 (개발용)
    class DummySupabase:
        def table(self, name):
            return DummyTable()
    
    class DummyTable:
        def select(self, *args):
            return self
        def eq(self, *args):
            return self
        def insert(self, data):
            return DummyResult()
        def update(self, data):
            return DummyResult()
        def delete(self):
            return self
        def execute(self):
            return DummyResult()
        def limit(self, n):
            return self
        def order(self, field, desc=False):
            return self
        def in_(self, field, values):
            return self
        def range(self, start, end):
            return self
    
    class DummyResult:
        def __init__(self):
            self.data = []
    
    supabase = DummySupabase()
    try:
        print("[OK] 더미 Supabase 클라이언트 초기화 완료")
    except Exception:
        pass
else:
    # SSL 인증서 검증 설정 (app.py와 동일)
    verify_ssl = os.getenv('SUPABASE_VERIFY_SSL', 'false').lower() in ('true', '1', 'yes')
    
    if not verify_ssl:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        import ssl
        ssl._create_default_https_context = ssl._create_unverified_context
        os.environ['PYTHONHTTPSVERIFY'] = '0'
        os.environ['CURL_CA_BUNDLE'] = ''
        os.environ['REQUESTS_CA_BUNDLE'] = ''
    
    try:
        # Supabase 클라이언트를 기본 방식으로 생성
        supabase: Client = create_client(supabase_url, supabase_key)
        
        # SSL 검증이 비활성화된 경우, 내부 httpx 클라이언트의 verify 옵션만 변경
        if not verify_ssl:
            try:
                if hasattr(supabase, 'postgrest') and hasattr(supabase.postgrest, 'session'):
                    original_client = supabase.postgrest.session
                    if hasattr(original_client, 'base_url'):
                        from httpx import Client as HttpxClient
                        new_client = HttpxClient(
                            base_url=original_client.base_url,
                            verify=False,
                            timeout=original_client.timeout if hasattr(original_client, 'timeout') else 30.0,
                            headers=original_client.headers if hasattr(original_client, 'headers') else {}
                        )
                        supabase.postgrest.session = new_client
            except Exception:
                pass
    except Exception as e:
        print(f"[ERROR] Supabase 클라이언트 생성 실패: {e}")
        import traceback
        traceback.print_exc()
        # 에러가 발생해도 서버는 시작되도록 더미 클라이언트 사용
        class DummySupabase:
            def table(self, name):
                return DummyTable()
        
        class DummyTable:
            def select(self, *args):
                return self
            def eq(self, *args):
                return self
            def insert(self, data):
                return DummyResult()
            def update(self, data):
                return DummyResult()
            def delete(self):
                return self
            def execute(self):
                return DummyResult()
            def limit(self, n):
                return self
            def order(self, field, desc=False):
                return self
            def in_(self, field, values):
                return self
            def range(self, start, end):
                return self
        
        class DummyResult:
            def __init__(self):
                self.data = []
        
        supabase = DummySupabase()
    
    try:
        print("[OK] Supabase 클라이언트 초기화 완료")
    except Exception:
        pass
    supabase_service: Client | None = None
    try:
        if supabase_service_key:
            # Supabase 서비스 클라이언트 생성
            supabase_service = create_client(supabase_url, supabase_service_key)
            
            # SSL 검증이 비활성화된 경우, 내부 httpx 클라이언트의 verify 옵션만 변경
            if not verify_ssl:
                try:
                    if hasattr(supabase_service, 'postgrest') and hasattr(supabase_service.postgrest, 'session'):
                        original_client = supabase_service.postgrest.session
                        if hasattr(original_client, 'base_url'):
                            from httpx import Client as HttpxClient
                            new_client = HttpxClient(
                                base_url=original_client.base_url,
                                verify=False,
                                timeout=original_client.timeout if hasattr(original_client, 'timeout') else 30.0,
                                headers=original_client.headers if hasattr(original_client, 'headers') else {}
                            )
                            supabase_service.postgrest.session = new_client
                except Exception:
                    pass
            try:
                print("[OK] Supabase 서비스 키 클라이언트 준비(스토리지 전용)")
            except Exception:
                pass
    except Exception:
        supabase_service = None
        try:
            print("[WARN] Supabase 서비스 키 클라이언트 초기화 실패: 환경 변수 또는 권한을 확인하세요")
        except Exception:
            pass

# JWT 토큰 검증 함수
def verify_token(token):
    try:
        payload = jwt.decode(token, str(SECRET_KEY or 'dev-secret-key-change-in-production'), algorithms=['HS256'])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None
@sites_bp.route('/admin/emergency-promote', methods=['POST'])
def emergency_promote():
    """비상 승격: 관리자 0명일 때에만 .env 코드로 1명 승격(1회성 권장)
    입력: { user_id, code }
    보호: 서버 환경변수 코드 일치 + 현재 활성 관리자 수 0명 조건
    """
    try:
        data = request.get_json() or {}
        code = (data.get('code') or '').strip()
        user_id = data.get('user_id')
        if not code or not user_id:
            return jsonify({'error': 'code와 user_id가 필요합니다.'}), 400
        if not EMERGENCY_ADMIN_CODE:
            return jsonify({'error': '비상승격 코드가 설정되어 있지 않습니다.'}), 403
        if code != EMERGENCY_ADMIN_CODE:
            return jsonify({'error': '비상승격 코드가 올바르지 않습니다.'}), 403

        try:
            rows = supabase.table('users').select('id, is_active, deleted_at').eq('user_role','admin').execute()
            admins = rows.data or []
            active_admins = [u for u in admins if (u.get('is_active') is not False) and (u.get('deleted_at') is None)]
            if len(active_admins) > 0:
                return jsonify({'error': '관리자가 이미 존재합니다. 비상승격은 관리자 0명일 때만 가능합니다.'}), 409
        except Exception:
            # 컬럼이 없으면 단순 카운트
            rows = supabase.table('users').select('id').eq('user_role','admin').execute()
            if len(rows.data or []) > 0:
                return jsonify({'error': '관리자가 이미 존재합니다.'}), 409

        # RPC 우선 호출
        try:
            if 'supabase_service' in globals() and supabase_service:
                rpc_res = supabase_service.rpc('promote_to_admin', {'p_user_id': user_id}).execute()
                return jsonify({'message': '비상 승격 완료', 'result': getattr(rpc_res, 'data', None)}), 200
        except Exception:
            pass

        # 폴백 업데이트
        res = supabase.table('users').update({'user_role':'admin', 'updated_at': datetime.utcnow().isoformat()}).eq('id', user_id).execute()
        return jsonify({'message': '비상 승격 완료(폴백)', 'user': (res.data[0] if res.data else None)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# =============================
# 관리자: 사용자 역할 변경
# =============================
@sites_bp.route('/admin/users/<int:user_id>', methods=['PATCH'])
def admin_update_user_role(user_id):
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        if payload.get('user_role') != 'admin':
            return jsonify({'error': '관리자만 접근 가능합니다.'}), 403

        body = request.get_json() or {}
        new_role = (body.get('user_role') or '').strip()
        if new_role not in ['admin', 'user']:
            return jsonify({'error': 'user_role은 admin 또는 user만 가능합니다.'}), 400
        # 자기 자신을 user로 강등 금지(옵션)
        if user_id == payload.get('user_id') and new_role != 'admin':
            return jsonify({'error': '자기 자신을 일반사용자로 강등할 수 없습니다.'}), 400

        # 관리자 승격 제한: 현재 활성 관리자 수 < 2 일 때만 허용
        if new_role == 'admin':
            try:
                # 우선 RPC 경로 시도(원자성 보장)
                if 'supabase_service' in globals() and supabase_service:
                    rpc_res = supabase_service.rpc('promote_to_admin', {'p_user_id': user_id}).execute()
                    return jsonify({'message': '관리자로 승격되었습니다.', 'result': getattr(rpc_res, 'data', None)}), 200
            except Exception as rpc_err:
                # RPC 실패 시 서버 측 폴백(경합 가능성 있지만 UX 보장)
                try:
                    rows = supabase.table('users').select('id, is_active, deleted_at').eq('user_role','admin').execute()
                    admins = rows.data or []
                    def _is_active(u):
                        return (u.get('is_active') is not False) and (u.get('deleted_at') is None)
                    active_admins = [u for u in admins if _is_active(u)]
                    if len(active_admins) >= 2:
                        return jsonify({'error': '관리자는 최대 2명입니다.'}), 409
                except Exception:
                    # is_active/deleted_at 컬럼이 없는 경우: 단순 카운트로 제한
                    rows = supabase.table('users').select('id').eq('user_role','admin').execute()
                    if len(rows.data or []) >= 2:
                        return jsonify({'error': '관리자는 최대 2명입니다.'}), 409

                res = supabase.table('users').update({'user_role': 'admin'}).eq('id', user_id).execute()
                return jsonify({'message': '관리자로 승격되었습니다.(폴백)', 'user': (res.data[0] if res.data else None)}), 200

        # 일반 사용자 강등 또는 기타 변경
        res = supabase.table('users').update({'user_role': new_role}).eq('id', user_id).execute()
        return jsonify({'message': '역할이 변경되었습니다.', 'user': (res.data[0] if res.data else None)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 사용자 목록 조회 API (연락처용)
@sites_bp.route('/users', methods=['GET'])
def get_users():
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401

        # 관리자 전용으로 제한
        if payload.get('user_role') != 'admin':
            return jsonify({'error': '관리자만 접근 가능합니다.'}), 403

        q = request.args.get('q')  # 검색어

        query = supabase.table('users').select('id, email, name, phone, user_role')
        rows = query.execute()
        items = rows.data or []

        # 더미 모드에서도 관리자 전용 정책 유지

        # 간단한 서버측 필터링 (name 포함 검색)
        if q:
            ql = q.lower()
            items = [it for it in items if (it.get('name','').lower().find(ql) >= 0)]

        return jsonify({'items': items}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 마스터 인명 조회 (역할별 필터 및 검색)
@sites_bp.route('/contacts-master', methods=['GET'])
def get_contacts_master():
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401

        role = request.args.get('role')  # pm | sales | None
        q = request.args.get('q')  # 검색어

        query = supabase.table('contacts_master').select('*').eq('active', True)
        if role in ['pm','sales']:
            query = query.eq('role', role)
        rows = query.execute()
        items = rows.data or []

        # 간단한 서버측 필터링 (name 포함 검색)
        if q:
            ql = q.lower()
            items = [it for it in items if (it.get('name','').lower().find(ql) >= 0)]

        return jsonify({'items': items}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 마스터 인명 추가/수정 (관리자 전용)
@sites_bp.route('/contacts-master', methods=['POST','PATCH'])
def upsert_contacts_master():
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        if payload.get('user_role') != 'admin':
            return jsonify({'error': '관리자만 접근 가능합니다.'}), 403

        data = request.get_json() or {}
        # 기대 필드: id(optional), name, role(pm|sales), phone, active
        item = {
            'name': data.get('name'),
            'role': data.get('role'),
            'phone': data.get('phone'),
            'active': data.get('active', True),
            'updated_at': datetime.utcnow().isoformat()
        }
        if not item['name'] or item['role'] not in ['pm','sales']:
            return jsonify({'error': 'name과 role(pm|sales)은 필수입니다.'}), 400

        if data.get('id'):
            # update
            res = supabase.table('contacts_master').update(item).eq('id', data['id']).execute()
        else:
            # insert
            item['created_at'] = datetime.utcnow().isoformat()
            res = supabase.table('contacts_master').insert(item).execute()

        return jsonify({'item': (res.data[0] if res.data else None)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 현장 등록
@sites_bp.route('/sites', methods=['POST'])
def create_site():
    try:
        # 인증 확인
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        
        data = request.get_json()
        
        # 필수 필드 검증
        required_fields = ['project_no', 'construction_company', 'site_name', 'address_sido', 'address_sigungu', 'household_count']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field}는 필수 입력 항목입니다.'}), 400
        
        # 등록번호는 더 이상 사용하지 않음
        
        # 현장 데이터 생성
        base_address = data.get('address') or ' '.join([part for part in [data.get('address_sido'), data.get('address_sigungu')] if part])
        if not base_address:
            return jsonify({'error': '주소 정보가 올바르지 않습니다.'}), 400

        site_data = {
            'project_no': data['project_no'],
            'construction_company': data['construction_company'],
            'site_name': data['site_name'],
            'address': base_address,
            'address_sido': data.get('address_sido'),
            'address_sigungu': data.get('address_sigungu'),
            'detail_address': data.get('detail_address', ''),
            'household_count': data['household_count'],
            'registration_date': data.get('registration_date'),
            'delivery_date': data.get('delivery_date'),
            'completion_date': data.get('completion_date'),
            'certification_audit': data.get('certification_audit', 'N'),
            'home_iot': data.get('home_iot', 'N'),
            'product_bi': data.get('product_bi'),
            'special_notes': (data.get('special_notes')[:1000] if data.get('special_notes') else None),
            'external_network_enabled': (data.get('external_network_enabled') or 'N'),
            'external_network_period': (data.get('external_network_period') if (data.get('external_network_enabled') == 'Y') else None),
            'created_by': payload['user_id'],
            'created_at': datetime.utcnow().isoformat()
        }
        
        result = supabase.table('sites').insert(site_data).execute()
        
        # 더미 데이터인 경우에도 성공으로 처리
        if result.data or not supabase_url or not supabase_key:
            # 더미 데이터인 경우 가짜 현장 데이터 반환
            dummy_site = {
                'id': 1,
                'project_no': site_data['project_no'],
                'construction_company': site_data['construction_company'],
                'site_name': site_data['site_name'],
                'address': site_data['address'],
                'created_by': site_data['created_by']
            }
            return jsonify({
                'message': '현장이 성공적으로 등록되었습니다.',
                'site': dummy_site if not result.data else result.data[0]
            }), 201
        else:
            return jsonify({'error': '현장 등록 중 오류가 발생했습니다.'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 현장 목록 조회
@sites_bp.route('/sites', methods=['GET'])
def get_sites():
    try:
        # 인증 확인
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        
        # 관리자는 모든 현장 조회, 일반사용자는 본인이 등록한 현장만 조회
        if payload['user_role'] == 'admin':
            sites = supabase.table('sites').select('*').order('id', desc=True).execute()
        else:
            sites = supabase.table('sites').select('*').eq('created_by', payload['user_id']).order('id', desc=True).execute()
        
        # 더미 데이터인 경우 빈 배열 반환
        sites_data = sites.data if sites.data else []
        
        return jsonify({'sites': sites_data}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 특정 현장 상세 조회
@sites_bp.route('/sites/<int:site_id>', methods=['GET'])
def get_site_detail(site_id):
    try:
        # 인증 확인
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        
        # 현장 조회
        site = supabase.table('sites').select('*').eq('id', site_id).execute()
        
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        
        site_info = site.data[0]
        
        # 권한 확인 (관리자가 아닌 경우 본인이 등록한 현장만 조회 가능)
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403
        
        return jsonify({'site': site_info}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 현장 기본정보 수정
@sites_bp.route('/sites/<int:site_id>', methods=['PATCH','PUT'])
def update_site(site_id):
    try:
        print(f"🔧 현장 수정 요청: ID {site_id}")
        print(f"📝 요청 데이터: {request.get_json()}")
        print(f"🔑 인증 헤더: {request.headers.get('Authorization', '없음')}")
        print(f"🌐 Supabase URL: {supabase_url}")
        print(f"🔑 Supabase Key: {supabase_key[:20]}..." if supabase_key else "❌ Supabase Key 없음")
        
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            print("❌ 인증 헤더 없음")
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        
        # 권한 확인
        print(f"🔍 권한 확인 중: site_id={site_id}")
        try:
            site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
            print(f"✅ 권한 확인 성공: {site.data}")
        except Exception as db_error:
            print(f"❌ 권한 확인 실패: {db_error}")
            return jsonify({'error': f'데이터베이스 연결 오류: {str(db_error)}'}), 500
            
        if not site.data:
            print("❌ 현장을 찾을 수 없음")
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            print("❌ 접근 권한 없음")
            return jsonify({'error': '접근 권한이 없습니다.'}), 403
        
        data = request.get_json()
        address_sido = data.get('address_sido')
        address_sigungu = data.get('address_sigungu')
        base_address = data.get('address')
        if not base_address:
            candidate_sido = address_sido if address_sido is not None else site_info.get('address_sido')
            candidate_sigungu = address_sigungu if address_sigungu is not None else site_info.get('address_sigungu')
            base_address = ' '.join([part for part in [candidate_sido, candidate_sigungu] if part])
        update_data = {
            'project_no': data.get('project_no'),
            'construction_company': data.get('construction_company'),
            'site_name': data.get('site_name'),
            'address': base_address,
            'address_sido': address_sido,
            'address_sigungu': address_sigungu,
            'detail_address': data.get('detail_address'),
            'household_count': data.get('household_count'),
            'registration_date': data.get('registration_date') if data.get('registration_date') else None,
            'delivery_date': data.get('delivery_date') if data.get('delivery_date') else None,
            'completion_date': data.get('completion_date') if data.get('completion_date') else None,
            'certification_audit': data.get('certification_audit'),
            'home_iot': data.get('home_iot'),
            'product_bi': data.get('product_bi'),
            'special_notes': (data.get('special_notes')[:1000] if data.get('special_notes') else None),
            'external_network_enabled': data.get('external_network_enabled'),
            'external_network_period': (data.get('external_network_period') if (data.get('external_network_enabled') == 'Y') else None),
            'updated_at': datetime.utcnow().isoformat()
        }
        
        # None 값 제거
        update_data = {k: v for k, v in update_data.items() if v is not None}
        print(f"📝 업데이트할 데이터: {update_data}")
        
        try:
            result = supabase.table('sites').update(update_data).eq('id', site_id).execute()
            print(f"✅ 데이터베이스 업데이트 성공: {result.data}")
        except Exception as update_error:
            print(f"❌ 데이터베이스 업데이트 실패: {update_error}")
            return jsonify({'error': f'데이터베이스 업데이트 오류: {str(update_error)}'}), 500
        
        if result.data:
            return jsonify({'message': '현장 정보가 수정되었습니다.', 'site': result.data[0]}), 200
        else:
            return jsonify({'error': '현장 정보 수정 중 오류가 발생했습니다.'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 현장 연락처 조회
@sites_bp.route('/sites/<int:site_id>/contacts', methods=['GET'])
def get_site_contacts(site_id):
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        # 사진 목록은 로그인한 사용자라면 모두 열람 가능(팀 공유 정책 없음)
        site_info = site.data[0]
        contacts = supabase.table('site_contacts').select('*').eq('site_id', site_id).limit(1).execute()
        base = contacts.data[0] if contacts.data else None

        # 추가 연락처(복수) 목록 로드: sales|construction|installer|network
        def _load_list(kind: str):
            try:
                rows = supabase.table('site_contact_people').select('*').eq('site_id', site_id).eq('person_type', kind).order('id', desc=True).execute()
                return [{'name': (r.get('name') or ''), 'phone': (r.get('phone') or '')} for r in (rows.data or [])]
            except Exception as e_list:
                msg = str(e_list)
                # 테이블이 없는 경우에도 빈 리스트 반환
                if 'site_contact_people' in msg and ('does not exist' in msg or 'relation' in msg or 'schema cache' in msg):
                    return []
                # 기타 오류는 빈 리스트로 처리(UX 우선)
                return []

        result = base or {}
        result = dict(result)
        result['sales_list'] = _load_list('sales')
        result['construction_list'] = _load_list('construction')
        result['installer_list'] = _load_list('installer')
        result['network_list'] = _load_list('network')
        return jsonify({'contacts': result}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 현장 제품수량 저장(업서트) - 프론트엔드용
@sites_bp.route('/sites/<int:site_id>/products', methods=['POST'])
def upsert_site_products(site_id):
    try:
        print(f"🔍 제품수량 저장 요청 - 현장 ID: {site_id}")
        print(f"📝 Raw 데이터: {request.get_data()}")
        print(f"📝 Content-Type: {request.headers.get('Content-Type', '없음')}")
        
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        
        # JSON 데이터 안전하게 파싱
        try:
            data = request.get_json()
            print(f"📝 파싱된 JSON 데이터: {data}")
        except Exception as json_error:
            print(f"❌ JSON 파싱 오류: {json_error}")
            return jsonify({'error': '잘못된 JSON 형식입니다.'}), 400
        
        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        # 사진 업로드는 로그인한 사용자라면 모두 가능(팀 공유 정책 없음)
        
        payload_data = {
            'site_id': site_id,
            'project_no': data.get('project_no'),
            'wallpad_model': data.get('wallpad_model'),
            'wallpad_qty': data.get('wallpad_qty', 0),
            'doorphone_model': data.get('doorphone_model'),
            'doorphone_qty': data.get('doorphone_qty', 0),
            'lobbyphone_model': data.get('lobbyphone_model'),
            'lobbyphone_qty': data.get('lobbyphone_qty', 0),
            'guardphone_model': data.get('guardphone_model'),
            'guardphone_qty': data.get('guardphone_qty', 0),
            'magnet_sensor_model': data.get('magnet_sensor_model'),
            'magnet_sensor_qty': data.get('magnet_sensor_qty', 0),
            'motion_sensor_model': data.get('motion_sensor_model'),
            'motion_sensor_qty': data.get('motion_sensor_qty', 0),
            'opener_model': data.get('opener_model'),
            'opener_qty': data.get('opener_qty', 0),
            'updated_at': datetime.utcnow().isoformat()
        }
        
        # None 값 제거
        payload_data = {k: v for k, v in payload_data.items() if v is not None}
        print(f"💾 저장할 데이터: {payload_data}")
        
        existing = supabase.table('site_products').select('id').eq('site_id', site_id).limit(1).execute()
        if existing.data:
            # 기존 데이터 업데이트
            result = supabase.table('site_products').update(payload_data).eq('id', existing.data[0]['id']).execute()
        else:
            # 새 데이터 삽입
            payload_data['created_at'] = datetime.utcnow().isoformat()
            result = supabase.table('site_products').insert(payload_data).execute()
        
        print(f"✅ 제품수량 저장 성공: {result.data[0] if result.data else 'None'}")
        if result.data:
            return jsonify({'message': '제품수량 정보가 저장되었습니다.', 'products': result.data[0]}), 200
        else:
            return jsonify({'error': '제품수량 정보 저장 중 오류가 발생했습니다.'}), 500
            
    except Exception as e:
        print(f"❌ 제품수량 저장 오류: {str(e)}")
        return jsonify({'error': str(e)}), 500

# 현장 연락처 저장(업서트)
@sites_bp.route('/sites/<int:site_id>/contacts', methods=['POST'])
def upsert_site_contacts(site_id):
    try:
        print(f"🔍 연락처 저장 요청 - 현장 ID: {site_id}")
        print(f"📝 Raw 데이터: {request.get_data()}")
        print(f"📝 Content-Type: {request.headers.get('Content-Type', '없음')}")
        
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        
        # JSON 데이터 안전하게 파싱
        try:
            data = request.get_json()
            print(f"📝 파싱된 JSON 데이터: {data}")
        except Exception as json_error:
            print(f"❌ JSON 파싱 오류: {json_error}")
            return jsonify({'error': '잘못된 JSON 형식입니다.'}), 400
        
        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403
        
        payload_data = {
            'site_id': site_id,
            'project_no': data.get('project_no'),
            'pm_name': data.get('pm_name'),
            'pm_phone': data.get('pm_phone'),
            # 단일 필드(하위 리스트의 첫 항목으로 보정 가능)
            'sales_manager_name': data.get('sales_manager_name'),
            'sales_manager_phone': data.get('sales_manager_phone'),
            'construction_manager_name': data.get('construction_manager_name'),
            'construction_manager_phone': data.get('construction_manager_phone'),
            'installer_name': data.get('installer_name'),
            'installer_phone': data.get('installer_phone'),
            'network_manager_name': data.get('network_manager_name'),
            'network_manager_phone': data.get('network_manager_phone'),
            'updated_at': datetime.utcnow().isoformat()
        }
        
        # None 값 제거
        payload_data = {k: v for k, v in payload_data.items() if v is not None}
        print(f"💾 저장할 데이터: {payload_data}")
        
        # 1) 메인 레코드 upsert
        existing = supabase.table('site_contacts').select('id').eq('site_id', site_id).limit(1).execute()
        if existing.data:
            contact_id = existing.data[0]['id']
            result = supabase.table('site_contacts').update(payload_data).eq('id', contact_id).execute()
        else:
            result = supabase.table('site_contacts').insert(payload_data).execute()

        # 2) 복수 연락처 리스트 저장(있다면 교체 방식)
        def _normalize_list(arr):
            if not isinstance(arr, list):
                return []
            norm = []
            for it in arr:
                name = str((it or {}).get('name') or '').strip()
                phone = str((it or {}).get('phone') or '').strip()
                if not name and not phone:
                    continue
                norm.append({'name': name, 'phone': phone})
            return norm

        sales_list = _normalize_list(data.get('sales_list'))
        construction_list = _normalize_list(data.get('construction_list'))
        installer_list = _normalize_list(data.get('installer_list'))
        network_list = _normalize_list(data.get('network_list'))

        # 단일 필드 보정: 첫 항목을 반영(이전 스키마와 호환)
        def _set_first_to_payload(list_val, name_key, phone_key):
            if list_val and not payload_data.get(name_key):
                payload_data[name_key] = list_val[0]['name']
            if list_val and not payload_data.get(phone_key):
                payload_data[phone_key] = list_val[0]['phone']
        _set_first_to_payload(sales_list, 'sales_manager_name', 'sales_manager_phone')
        _set_first_to_payload(construction_list, 'construction_manager_name', 'construction_manager_phone')
        _set_first_to_payload(installer_list, 'installer_name', 'installer_phone')
        _set_first_to_payload(network_list, 'network_manager_name', 'network_manager_phone')

        # 테이블 없을 수 있으므로 안전 처리
        def _replace(kind: str, items: list):
            try:
                # 기존 삭제
                supabase.table('site_contact_people').delete().eq('site_id', site_id).eq('person_type', kind).execute()
            except Exception as e_del:
                # 생성 안된 경우 무시
                if 'site_contact_people' not in str(e_del):
                    pass
            if not items:
                return
            try:
                payload_rows = [{
                    'site_id': site_id,
                    'person_type': kind,
                    'name': it['name'],
                    'phone': it['phone'],
                    'created_by': payload['user_id'],
                    'created_at': datetime.utcnow().isoformat(),
                    'updated_at': datetime.utcnow().isoformat()
                } for it in items]
                supabase.table('site_contact_people').insert(payload_rows).execute()
            except Exception as e_ins:
                # 테이블이 없으면 조용히 패스(프론트에서 SQL 적용 유도)
                if 'site_contact_people' not in str(e_ins):
                    print(f"⚠️ site_contact_people 저장 오류({kind}): {e_ins}")

        _replace('sales', sales_list)
        _replace('construction', construction_list)
        _replace('installer', installer_list)
        _replace('network', network_list)

        print(f"✅ 연락처 저장 성공: {result.data[0] if result.data else 'None'}")
        return jsonify({'message': '연락처가 저장되었습니다.', 'contacts': result.data[0] if result.data else payload_data}), 200
    except Exception as e:
        print(f"❌ 연락처 저장 오류: {str(e)}")
        return jsonify({'error': str(e)}), 500

# 세대부연동 조회 (조명SW/대기전력SW/가스감지기/VPN/일괄소등 등)
@sites_bp.route('/sites/<int:site_id>/integrations/household', methods=['GET'])
def get_household_integrations(site_id):
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403

        types = ['lighting_sw','standby_power_sw','gas_detector','heating','ventilation','door_lock','air_conditioner','real_time_metering','environment_sensor','vpn','all_off_switch','bathroom_phone','kitchen_tv']
        rows = supabase.table('site_household_integrations').select('*').eq('site_id', site_id).in_('integration_type', types).execute()
        return jsonify({'items': rows.data or []}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 세대부연동 저장(업서트)
@sites_bp.route('/sites/<int:site_id>/integrations/household', methods=['POST'])
def upsert_household_integrations(site_id):
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403

        data = request.get_json() or {}
        items = data.get('items', [])
        print(f"📝 세대부 저장 요청 items: {items}")

        def _normalize(v):
            if v is None:
                return None
            if isinstance(v, str):
                v2 = v.strip()
                return v2 if v2 != '' else None
            return v

        def _yn(v):
            return 'Y' if str(v or 'N').strip().upper() == 'Y' else 'N'

        saved = []
        allowed = ['lighting_sw','standby_power_sw','gas_detector','heating','ventilation','door_lock','air_conditioner','real_time_metering','environment_sensor','vpn','all_off_switch','bathroom_phone','kitchen_tv']
        for item in items:
            itype = (item.get('integration_type') or '').strip()
            if itype not in allowed:
                print(f"⚠️ 허용되지 않은 타입(세대부): {itype}")
                continue
            
            # 저장할 의미 있는 데이터가 있는지 확인
            enabled = _yn(item.get('enabled'))
            project_no = _normalize(item.get('project_no'))
            company_name = _normalize(item.get('company_name'))
            contact_person = _normalize(item.get('contact_person'))
            contact_phone = _normalize(item.get('contact_phone'))
            notes = _normalize(item.get('notes'))
            
            # enabled가 'N'이고 다른 모든 필드가 비어있으면 저장하지 않음
            has_data = enabled == 'Y' or project_no or company_name or contact_person or contact_phone or notes
            if not has_data:
                print(f"⏭️ 저장할 데이터 없음(세대부): {itype} - 모든 필드가 비어있음")
                continue
            
            payload_data = {
                'site_id': site_id,
                'project_no': project_no,
                'integration_type': itype,
                'enabled': enabled,
                'company_name': company_name,
                'contact_person': contact_person,
                'contact_phone': contact_phone,
                'notes': notes,
                'updated_at': datetime.utcnow().isoformat()
            }
            print(f"➡️ 업서트 시도(세대부): {payload_data}")

            # 1) 업데이트 우선(site_id + integration_type)
            try:
                upd = supabase.table('site_household_integrations').update(payload_data).eq('site_id', site_id).eq('integration_type', itype).execute()
                if upd.data:
                    print(f"✅ 업데이트 성공(세대부): {upd.data}")
                    saved.append(upd.data[0])
                    continue
            except Exception as e_upd:
                print(f"❌ 업데이트 오류(세대부): {str(e_upd)}")

            # 2) 없으면 삽입
            try:
                payload_insert = dict(payload_data)
                payload_insert['created_at'] = datetime.utcnow().isoformat()
                ins = supabase.table('site_household_integrations').insert(payload_insert).execute()
                print(f"✅ 삽입 성공(세대부): {ins.data}")
                if ins.data:
                    saved.append(ins.data[0])
            except Exception as e_ins:
                # 삽입 실패 시 해당 항목만 건너뛰고 계속 진행
                print(f"⚠️ 삽입 실패(세대부) - 항목 건너뜀: {itype}, 오류: {str(e_ins)}")
                # 에러가 발생해도 다른 항목 처리를 계속함
                continue

        # 저장된 항목이 없을 때 안내 메시지 반환
        if not saved:
            return jsonify({'message': '저장할 내용이 없습니다.', 'items': [], 'no_data': True}), 200

        return jsonify({'message': '세대부연동이 저장되었습니다.', 'items': saved}), 200
    except Exception as e:
        print(f"❌ 세대부연동 전체 오류: {str(e)}")
        return jsonify({'error': '세대부연동 저장 실패', 'error_detail': str(e)}), 500

# 공용부연동 조회 (주차관제/원격검침/CCTV)
@sites_bp.route('/sites/<int:site_id>/integrations/common', methods=['GET'])
def get_common_integrations(site_id):
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403

        types = ['parking_control','remote_metering','cctv','elevator','parcel','ev_charger','parking_location','onepass','rf_card']
        rows = supabase.table('site_common_integrations').select('*').eq('site_id', site_id).in_('integration_type', types).execute()
        return jsonify({'items': rows.data or []}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 현장 세대부연동 저장(업서트) - 프론트엔드용
@sites_bp.route('/sites/<int:site_id>/household', methods=['POST'])
def upsert_site_household(site_id):
    try:
        print(f"🔍 세대부연동 저장 요청 - 현장 ID: {site_id}")
        print(f"📝 Raw 데이터: {request.get_data()}")
        print(f"📝 Content-Type: {request.headers.get('Content-Type', '없음')}")
        
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        
        # JSON 데이터 안전하게 파싱
        try:
            data = request.get_json()
            print(f"📝 파싱된 JSON 데이터: {data}")
        except Exception as json_error:
            print(f"❌ JSON 파싱 오류: {json_error}")
            return jsonify({'error': '잘못된 JSON 형식입니다.'}), 400
        
        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403
        
        payload_data = {
            'site_id': site_id,
            'project_no': data.get('project_no'),
            'lighting_enabled': data.get('lighting_enabled', 'N'),
            'lighting_company': data.get('lighting_company'),
            'standby_enabled': data.get('standby_enabled', 'N'),
            'standby_company': data.get('standby_company'),
            'gas_enabled': data.get('gas_enabled', 'N'),
            'gas_company': data.get('gas_company'),
            'updated_at': datetime.utcnow().isoformat()
        }
        
        # None 값 제거
        payload_data = {k: v for k, v in payload_data.items() if v is not None}
        print(f"💾 저장할 데이터: {payload_data}")
        
        existing = supabase.table('site_household_integrations').select('id').eq('site_id', site_id).limit(1).execute()
        if existing.data:
            # 기존 데이터 업데이트
            result = supabase.table('site_household_integrations').update(payload_data).eq('id', existing.data[0]['id']).execute()
        else:
            # 새 데이터 삽입
            payload_data['created_at'] = datetime.utcnow().isoformat()
            result = supabase.table('site_household_integrations').insert(payload_data).execute()
        
        print(f"✅ 세대부연동 저장 성공: {result.data[0] if result.data else 'None'}")
        if result.data:
            return jsonify({'message': '세대부연동 정보가 저장되었습니다.', 'household': result.data[0]}), 200
        else:
            return jsonify({'error': '세대부연동 정보 저장 중 오류가 발생했습니다.'}), 500
            
    except Exception as e:
        print(f"❌ 세대부연동 저장 오류: {str(e)}")
        return jsonify({'error': str(e)}), 500

# 현장 공용부연동 저장(업서트) - 프론트엔드용
@sites_bp.route('/sites/<int:site_id>/common', methods=['POST'])
def upsert_site_common(site_id):
    try:
        print(f"🔍 공용부연동 저장 요청 - 현장 ID: {site_id}")
        print(f"📝 Raw 데이터: {request.get_data()}")
        print(f"📝 Content-Type: {request.headers.get('Content-Type', '없음')}")
        
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        
        # JSON 데이터 안전하게 파싱
        try:
            data = request.get_json()
            print(f"📝 파싱된 JSON 데이터: {data}")
        except Exception as json_error:
            print(f"❌ JSON 파싱 오류: {json_error}")
            return jsonify({'error': '잘못된 JSON 형식입니다.'}), 400
        
        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403
        
        payload_data = {
            'site_id': site_id,
            'project_no': data.get('project_no'),
            'parking_enabled': data.get('parking_enabled', 'N'),
            'parking_company': data.get('parking_company'),
            'metering_enabled': data.get('metering_enabled', 'N'),
            'metering_company': data.get('metering_company'),
            'cctv_enabled': data.get('cctv_enabled', 'N'),
            'cctv_company': data.get('cctv_company'),
            'updated_at': datetime.utcnow().isoformat()
        }
        
        # None 값 제거
        payload_data = {k: v for k, v in payload_data.items() if v is not None}
        print(f"💾 저장할 데이터: {payload_data}")
        
        existing = supabase.table('site_common_integrations').select('id').eq('site_id', site_id).limit(1).execute()
        if existing.data:
            # 기존 데이터 업데이트
            result = supabase.table('site_common_integrations').update(payload_data).eq('id', existing.data[0]['id']).execute()
        else:
            # 새 데이터 삽입
            payload_data['created_at'] = datetime.utcnow().isoformat()
            result = supabase.table('site_common_integrations').insert(payload_data).execute()
        
        print(f"✅ 공용부연동 저장 성공: {result.data[0] if result.data else 'None'}")
        if result.data:
            return jsonify({'message': '공용부연동 정보가 저장되었습니다.', 'common': result.data[0]}), 200
        else:
            return jsonify({'error': '공용부연동 정보 저장 중 오류가 발생했습니다.'}), 500
            
    except Exception as e:
        print(f"❌ 공용부연동 저장 오류: {str(e)}")
        return jsonify({'error': str(e)}), 500

# 공용부연동 저장(업서트)
@sites_bp.route('/sites/<int:site_id>/integrations/common', methods=['POST'])
def upsert_common_integrations(site_id):
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403

        data = request.get_json() or {}
        items = data.get('items', [])
        print(f"📝 공용부 저장 요청 items: {items}")

        def _normalize(v):
            if v is None:
                return None
            if isinstance(v, str):
                v2 = v.strip()
                return v2 if v2 != '' else None
            return v

        def _yn(v):
            return 'Y' if str(v or 'N').strip().upper() == 'Y' else 'N'

        saved = []
        allowed = ['parking_control','remote_metering','cctv','elevator','parcel','ev_charger','parking_location','onepass','rf_card']
        for item in items:
            itype = (item.get('integration_type') or '').strip()
            if itype not in allowed:
                print(f"⚠️ 허용되지 않은 타입(공용부): {itype}")
                continue
            
            # 저장할 의미 있는 데이터가 있는지 확인
            enabled = _yn(item.get('enabled'))
            project_no = _normalize(item.get('project_no'))
            company_name = _normalize(item.get('company_name'))
            contact_person = _normalize(item.get('contact_person'))
            contact_phone = _normalize(item.get('contact_phone'))
            notes = _normalize(item.get('notes'))
            
            # enabled가 'N'이고 다른 모든 필드가 비어있으면 저장하지 않음
            has_data = enabled == 'Y' or project_no or company_name or contact_person or contact_phone or notes
            if not has_data:
                print(f"⏭️ 저장할 데이터 없음(공용부): {itype} - 모든 필드가 비어있음")
                continue
            
            payload_data = {
                'site_id': site_id,
                'project_no': project_no,
                'integration_type': itype,
                'enabled': enabled,
                'company_name': company_name,
                'contact_person': contact_person,
                'contact_phone': contact_phone,
                'notes': notes,
                'updated_at': datetime.utcnow().isoformat()
            }
            print(f"➡️ 업서트 시도(공용부): {payload_data}")

            # 1) 업데이트 우선(site_id + integration_type)
            try:
                upd = supabase.table('site_common_integrations').update(payload_data).eq('site_id', site_id).eq('integration_type', itype).execute()
                if upd.data:
                    print(f"✅ 업데이트 성공(공용부): {upd.data}")
                    saved.append(upd.data[0])
                    continue
            except Exception as e_upd:
                print(f"❌ 업데이트 오류(공용부): {str(e_upd)}")

            # 2) 없으면 삽입
            try:
                payload_insert = dict(payload_data)
                payload_insert['created_at'] = datetime.utcnow().isoformat()
                ins = supabase.table('site_common_integrations').insert(payload_insert).execute()
                print(f"✅ 삽입 성공(공용부): {ins.data}")
                if ins.data:
                    saved.append(ins.data[0])
            except Exception as e_ins:
                # 삽입 실패 시 해당 항목만 건너뛰고 계속 진행
                print(f"⚠️ 삽입 실패(공용부) - 항목 건너뜀: {itype}, 오류: {str(e_ins)}")
                # 에러가 발생해도 다른 항목 처리를 계속함
                continue

        # 저장된 항목이 없을 때 안내 메시지 반환
        if not saved:
            return jsonify({'message': '저장할 내용이 없습니다.', 'items': [], 'no_data': True}), 200

        return jsonify({'message': '공용부연동이 저장되었습니다.', 'items': saved}), 200
    except Exception as e:
        print(f"❌ 공용부연동 전체 오류: {str(e)}")
        return jsonify({'error': '공용부연동 저장 실패', 'error_detail': str(e)}), 500

# 제품수량 조회 (평면 스키마: wallpad_*, doorphone_*, lobbyphone_*, guardphone_*)
@sites_bp.route('/sites/<int:site_id>/products', methods=['GET'])
def get_site_products(site_id):
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403
        row = supabase.table('site_products').select('*').eq('site_id', site_id).limit(1).execute()
        return jsonify({'products': (row.data[0] if row.data else None)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# =============================
# 현장 사진등록 및 관리
# =============================

@sites_bp.route('/sites/<int:site_id>/photos', methods=['GET'])
def list_site_photos(site_id):
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401

        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403

        # 페이징 파라미터 (기본: page=1, page_size=20)
        try:
            page = max(1, int(request.args.get('page', '1')))
        except Exception:
            page = 1
        try:
            page_size = int(request.args.get('page_size', '20'))
            if page_size <= 0 or page_size > 100:
                page_size = 20
        except Exception:
            page_size = 20
        start = (page - 1) * page_size
        end = start + page_size - 1

        # count 포함하여 조회(가능한 경우)
        try:
            q = supabase.table('site_photos').select('*', count='exact').eq('site_id', site_id)
            # 소프트 삭제 제외(컬럼이 존재할 때만) - deleted_at이 null인 것만 조회
            try:
                q = q.is_('deleted_at', 'null')  # None 대신 'null' 문자열 사용
            except Exception:
                pass
            rows = q.order('id', desc=True).range(start, end).execute()
            total = getattr(rows, 'count', None)
        except Exception as e_sel:
            # 테이블 미생성/스키마 캐시 오류 시 빈 목록
            msg = str(e_sel)
            if 'site_photos' in msg and (
                'relation' in msg or 'does not exist' in msg or 'schema cache' in msg or 'PGRST' in msg
            ):
                return jsonify({'items': [], 'page': page, 'page_size': page_size, 'total': 0, 'has_more': False}), 200
            try:
                q2 = supabase.table('site_photos').select('*').eq('site_id', site_id)
                try:
                    q2 = q2.is_('deleted_at', 'null')  # None 대신 'null' 문자열 사용
                except Exception:
                    pass
                rows = q2.order('id', desc=True).range(start, end).execute()
                total = None
            except Exception as e_sel2:
                return jsonify({'error': f'사진 목록 조회 실패: {str(e_sel2)}'}), 500

        items = rows.data or []
        has_more = False
        if total is not None:
            has_more = (start + len(items)) < total
        else:
            has_more = len(items) == page_size

        return jsonify({'items': items, 'page': page, 'page_size': page_size, 'total': total, 'has_more': has_more}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@sites_bp.route('/sites/<int:site_id>/photos', methods=['POST'])
def upload_site_photo(site_id):
    """멀티파트 업로드: title(텍스트), file(이미지)
    - 촬영/앨범 모두 클라이언트가 파일로 업로드
    - 서버는 저장 시 uploaded_at(UTC ISO) 자동 기록
    - 파일은 backend/uploads/YYYY/MM/site_{site_id}_<timestamp>.<ext>
    - DB에는 파일 메타와 표시용 경로('/uploads/..') 저장
    """
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401

        # 권한 확인
        site = supabase.table('sites').select('id, created_by, site_name').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403

        # 멀티파트 파싱
        title = (request.form.get('title') or '').strip()
        file = request.files.get('file')
        if not file:
            return jsonify({'error': '이미지 파일이 필요합니다.'}), 400

        # 파일 크기 제한 (8MB)
        try:
            content = file.read()
        except Exception:
            return jsonify({'error': '파일을 읽을 수 없습니다.'}), 400
        MAX_SIZE = 8 * 1024 * 1024
        if content is None or len(content) == 0:
            return jsonify({'error': '빈 파일은 업로드할 수 없습니다.'}), 400
        if len(content) > MAX_SIZE:
            return jsonify({'error': '파일이 너무 큽니다. 최대 8MB까지 업로드할 수 있습니다.'}), 413

        now = datetime.utcnow()
        yyyy = str(now.year)
        mm = str(now.month).zfill(2)

        public_path = None
        # Supabase Storage 사용 여부
        if supabase_url and supabase_key:
            try:
                from werkzeug.utils import secure_filename
                orig = secure_filename(file.filename or 'image')
                ext = (orig.rsplit('.', 1)[-1].lower() if '.' in orig else 'jpg')
                object_path = f"site_{site_id}/{yyyy}/{mm}/site_{site_id}_{int(now.timestamp()*1000)}.{ext}"
                bucket = 'site-photos'

                # 업로드
                storage_client = None
                try:
                    storage_client = (supabase_service if 'supabase_service' in globals() and supabase_service else supabase)
                except Exception:
                    storage_client = supabase
                storage = storage_client.storage.from_(bucket)
                content_type = file.mimetype or 'application/octet-stream'
                # supabase-py는 file_options의 키를 camelCase로 기대합니다.
                storage.upload(object_path, content, { 'contentType': content_type, 'upsert': 'false' })

                # 퍼블릭 URL 구성
                public_path = f"{supabase_url}/storage/v1/object/public/{bucket}/{object_path}"
            except Exception as up_err:
                return jsonify({'error': '스토리지 업로드 실패', 'error_detail': str(up_err)}), 500
        else:
            # 로컬 저장 (더미 모드)
            base_dir = Path(__file__).resolve().parent
            uploads_dir = base_dir / 'uploads' / yyyy / mm
            uploads_dir.mkdir(parents=True, exist_ok=True)
            from werkzeug.utils import secure_filename
            orig = secure_filename(file.filename or 'image')
            ext = (orig.rsplit('.', 1)[-1].lower() if '.' in orig else 'jpg')
            fname = f"site_{site_id}_{int(now.timestamp()*1000)}.{ext}"
            full_path = uploads_dir / fname
            try:
                with open(full_path, 'wb') as f:
                    f.write(content)
            except Exception as werr:
                return jsonify({'error': '로컬 파일 저장 실패', 'error_detail': str(werr)}), 500
            public_path = f"/uploads/{yyyy}/{mm}/{fname}"

        row = {
            'site_id': site_id,
            'title': title or None,
            'image_url': public_path,
            'uploaded_at': now.isoformat(),
            'created_by': payload['user_id']
        }

        try:
            res = supabase.table('site_photos').insert(row).execute()
            saved = res.data[0] if res.data else row
        except Exception as ins_err:
            msg = str(ins_err)
            if 'site_photos' in msg and (
                'relation' in msg or 'does not exist' in msg or 'schema cache' in msg or 'PGRST' in msg
            ):
                return jsonify({'error': 'site_photos 테이블이 없습니다. Supabase SQL로 테이블을 먼저 생성해 주세요.'}), 500
            return jsonify({'error': '사진 메타 저장 실패', 'error_detail': msg}), 500

        return jsonify({'message': '사진이 저장되었습니다.', 'photo': saved}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@sites_bp.route('/sites/<int:site_id>/photos/<int:photo_id>', methods=['DELETE'])
def delete_site_photo(site_id, photo_id):
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401

        # 권한 확인: 사진 레코드와 현장 소유자 검사
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]

        photo_rows = supabase.table('site_photos').select('id, site_id, created_by, image_url').eq('id', photo_id).eq('site_id', site_id).limit(1).execute()
        if not photo_rows.data:
            return jsonify({'error': '사진을 찾을 수 없습니다.'}), 404
        photo = photo_rows.data[0]

        # 사진 삭제는 로그인한 사용자라면 모두 가능(팀 공유 정책 없음)

        # 관리자=하드 삭제, 일반=소프트 삭제
        hard_delete = (payload.get('user_role') == 'admin')
        if not hard_delete:
            # 소프트 삭제: deleted_at만 표시
            try:
                supabase.table('site_photos').update({'deleted_at': datetime.utcnow().isoformat()}).eq('id', photo_id).eq('site_id', site_id).execute()
                return jsonify({'message': '사진이 삭제되었습니다.(소프트)'}), 200
            except Exception:
                # 컬럼이 없으면 하드 삭제로 폴백
                pass

        # 파일 삭제 시도 (베스트에포트)
        try:
            public_path = photo.get('image_url') or ''
            if supabase_url and supabase_key and '/storage/v1/object/public/' in public_path:
                # 예: https://<proj>.supabase.co/storage/v1/object/public/site-photos/site_1/....jpg
                try:
                    bucket = 'site-photos'
                    prefix = f"{supabase_url}/storage/v1/object/public/{bucket}/"
                    if public_path.startswith(prefix):
                        object_path = public_path[len(prefix):]
                        storage_client = None
                        try:
                            storage_client = (supabase_service if 'supabase_service' in globals() and supabase_service else supabase)
                        except Exception:
                            storage_client = supabase
                        storage_client.storage.from_(bucket).remove([object_path])
                except Exception:
                    pass
            elif public_path.startswith('/uploads/'):
                # 로컬 파일 삭제
                rel = public_path[len('/uploads/'):]
                base_dir = Path(__file__).resolve().parent
                full_path = base_dir / 'uploads' / rel
                if full_path.exists():
                    try:
                        full_path.unlink(missing_ok=True)
                    except Exception:
                        pass
        except Exception:
            pass

        supabase.table('site_photos').delete().eq('id', photo_id).eq('site_id', site_id).execute()
        return jsonify({'message': '사진이 삭제되었습니다.(하드)'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# =============================
# 데이터 내보내기(관리자: 전체, 일반: 본인 현장)
# =============================
@sites_bp.route('/export', methods=['GET'])
def export_data():
    try:
        # 인증
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401

        user_id = payload.get('user_id')
        user_role = payload.get('user_role')

        # 파라미터
        fmt = (request.args.get('format') or 'both').lower()  # csv|xlsx|both
        scope = (request.args.get('scope') or 'auto').lower()  # auto|site
        site_id_param = request.args.get('site_id')
        include_photos = str(request.args.get('include_photos', 'true')).lower() in ['1','true','yes','y']
        start_date = (request.args.get('start_date') or '').strip()  # YYYY-MM-DD
        end_date = (request.args.get('end_date') or '').strip()      # YYYY-MM-DD

        # 접근 범위: 관리자면 전체, 일반이면 본인이 만든 현장만
        if user_role == 'admin':
            base_q = supabase.table('sites').select('id')
        else:
            base_q = supabase.table('sites').select('id').eq('created_by', user_id)

        if scope == 'site' and site_id_param:
            try:
                sid = int(site_id_param)
                sites_rows = base_q.eq('id', sid).order('id', desc=True).execute()
            except Exception:
                sites_rows = base_q.order('id', desc=True).execute()
        else:
            sites_rows = base_q.order('id', desc=True).execute()
        site_ids = [r['id'] for r in (sites_rows.data or [])]

        # 선택된 현장이 없으면 빈 ZIP 반환
        if not site_ids:
            buf = BytesIO()
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.writestr('README.txt', 'No data for export.')
            buf.seek(0)
            ts = datetime.utcnow().strftime('%Y%m%d_%H%M')
            return send_file(buf, mimetype='application/zip', as_attachment=True, download_name=f'export_{ts}.zip')

        def fetch_table(name, filter_by_site=True):
            q = supabase.table(name).select('*')
            if filter_by_site:
                q = q.in_('site_id', site_ids)
            rows = q.execute()
            return rows.data or []

        # 데이터 수집
        data_sites = fetch_table('sites', filter_by_site=False)
        # sites 범위를 사용자 범위로 축소(일반 사용자일 때)
        if user_role != 'admin':
            data_sites = [r for r in data_sites if r.get('id') in site_ids]
        data_contacts = fetch_table('site_contacts')
        # 복수 연락처 테이블은 없을 수 있으므로 예외 보호
        try:
            data_contact_people = fetch_table('site_contact_people')
        except Exception:
            data_contact_people = []
        data_products = fetch_table('site_products')
        data_work_items = fetch_table('work_items')
        # 소프트 삭제 제외 - deleted_at이 null인 것만 조회
        try:
            data_photos = supabase.table('site_photos').select('*').in_('site_id', site_ids).is_('deleted_at', 'null').execute().data or []
        except Exception:
            data_photos = fetch_table('site_photos')
        # 세대부 연동 및 공용부 연동 데이터
        try:
            data_household_integrations = fetch_table('site_household_integrations')
        except Exception:
            data_household_integrations = []
        try:
            data_common_integrations = fetch_table('site_common_integrations')
        except Exception:
            data_common_integrations = []

        # Excel 단일 시트용 병합 데이터프레임(table 구분 컬럼 포함)
        def df_with_table(rows, table_name):
            try:
                df = pd.DataFrame(rows)
            except Exception:
                df = pd.DataFrame()
            if 'table' not in df.columns:
                df['table'] = table_name
            else:
                df['table'] = table_name
            return df

        df_all = pd.concat([
            df_with_table(data_sites, 'sites'),
            df_with_table(data_contacts, 'site_contacts'),
            df_with_table(data_contact_people, 'site_contact_people'),
            df_with_table(data_products, 'site_products'),
            df_with_table(data_work_items, 'work_items'),
            df_with_table(data_photos, 'site_photos'),
            df_with_table(data_household_integrations, 'site_household_integrations'),
            df_with_table(data_common_integrations, 'site_common_integrations'),
        ], ignore_index=True, sort=False)

        # 컬럼명을 한국어로 변경 (실제 입력 항목명과 동일하게)
        column_mapping = {
            # sites 테이블
            'id': 'ID',
            'site_id': '현장ID',
            'project_no': '프로젝트 No.',
            'construction_company': '건설사',
            'site_name': '현장명',
            'address': '주소',
            'address_sido': '주소(시/도)',
            'address_sigungu': '주소(시/군/구)',
            'detail_address': '상세주소',
            'household_count': '세대수',
            'registration_date': '등록일',
            'delivery_date': '납품예정',
            'completion_date': '준공일',
            'certification_audit': '인증심사여부',
            'home_iot': '홈IoT연동여부',
            'product_bi': '제품 BI',
            'special_notes': '현장 특이사항',
            'external_network_enabled': '외부망 연동',
            'external_network_period': '가입기간',
            'created_by': '생성자ID',
            'created_at': '생성일시',
            'updated_at': '수정일시',
            # site_contacts 테이블
            'pm_name': 'PM 이름',
            'pm_phone': 'PM 전화번호',
            'sales_manager_name': '영업담당자',
            'sales_manager_phone': '영업담당자 전화',
            'construction_manager_name': '건설사 담당자',
            'construction_manager_phone': '건설사 담당자 전화',
            'installer_name': '설치점',
            'installer_phone': '설치점 전화',
            'network_manager_name': '네트워크점',
            'network_manager_phone': '네트워크점 전화',
            # site_products 테이블
            'product_type': '제품유형',
            'product_model': '제품모델',
            'quantity': '수량',
            'wallpad_model': '월패드 모델',
            'wallpad_qty': '월패드 수량',
            'doorphone_model': '도어폰 모델',
            'doorphone_qty': '도어폰 수량',
            'lobbyphone_model': '로비폰 모델',
            'lobbyphone_qty': '로비폰 수량',
            'guardphone_model': '경비실기 모델',
            'guardphone_qty': '경비실기 수량',
            'magnet_sensor_model': '자석감지기 모델',
            'magnet_sensor_qty': '자석감지기 수량',
            'motion_sensor_model': '동체감지기 모델',
            'motion_sensor_qty': '동체감지기 수량',
            'opener_model': '개폐기 모델',
            'opener_qty': '개폐기 수량',
            # site_household_integrations 테이블
            'integration_type': '연동유형',
            'enabled': '연동여부',
            'company_name': '업체명',
            'contact_person': '업체 담당자',
            'contact_phone': '연락처',
            'notes': '기타',
            # site_common_integrations 테이블 (동일한 컬럼명 사용)
            # work_items 테이블
            'content': '업무내용',
            'status': '상태',
            'alarm_date': '알람일자',
            'alarm_confirmed': '알람확인',
            'done_date': '완료일자',
            'work_type': '업무유형',
            'work_date': '업무일자',
            'worker_name': '작업자명',
            'work_content': '업무내용',
            'weather': '날씨',
            'temperature': '온도',
            # site_photos 테이블
            'title': '사진제목',
            'image_url': '사진URL',
            'photo_url': '사진URL',
            'photo_description': '사진설명',
            'uploaded_at': '업로드일시',
            'deleted_at': '삭제일시',
            # site_contact_people 테이블
            'name': '이름',
            'phone': '연락처',
            'contact_no': '연락처',
            'role': '역할',
            # 기타
            'table': '테이블',
            'start_date': '시작일',
            'end_date': '종료일',
            'site_manager': '현장관리자',
        }
        
        # df_all의 컬럼명 변경 (존재하는 컬럼만)
        existing_columns = {k: v for k, v in column_mapping.items() if k in df_all.columns}
        if existing_columns:
            df_all = df_all.rename(columns=existing_columns)
        
        # integration_type 값 한국어 변환
        integration_type_mapping = {
            # 세대부 연동
            'lighting_sw': '조명 SW 연동',
            'standby_power_sw': '대기전력 SW 연동',
            'gas_detector': '가스감지기 연동',
            'heating': '난방 연동',
            'ventilation': '환기 연동',
            'door_lock': '도어락 연동',
            'air_conditioner': '에어컨 연동',
            'real_time_metering': '실시간검침 연동',
            'environment_sensor': '환경감지 연동',
            'vpn': 'VPN 연동',
            'all_off_switch': '일괄소등스위치 연동',
            'bathroom_phone': '욕실폰 연동',
            'kitchen_tv': '주방 TV 연동',
            # 공용부 연동
            'parking_control': '주차관제 연동',
            'remote_metering': '원격검침 연동',
            'cctv': 'CCTV 연동',
            'elevator': '엘리베이터 연동',
            'parcel': '무인택배 연동',
            'ev_charger': '전기차충전 연동',
            'parking_location': '주차위치 연동',
            'onepass': '원패스 연동',
            'rf_card': 'RF 카드 연동',
        }
        
        # integration_type 컬럼이 있으면 값 변환
        if '연동유형' in df_all.columns:
            df_all['연동유형'] = df_all['연동유형'].map(lambda x: integration_type_mapping.get(x, x) if pd.notna(x) else x)
        elif 'integration_type' in df_all.columns:
            df_all['integration_type'] = df_all['integration_type'].map(lambda x: integration_type_mapping.get(x, x) if pd.notna(x) else x)

        # ZIP 빌드
        ts = datetime.utcnow().strftime('%Y%m%d_%H%M')
        buf = BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            # CSV들
            def write_csv(path, rows):
                try:
                    import csv
                    from io import StringIO
                    sio = StringIO()
                    if rows:
                        cols = sorted({k for r in rows for k in r.keys()})
                    else:
                        cols = []
                    
                    # 컬럼명을 한국어로 변경
                    korean_cols = [column_mapping.get(col, col) for col in cols]
                    
                    writer = csv.DictWriter(sio, fieldnames=korean_cols, extrasaction='ignore')
                    writer.writeheader()
                    for r in rows:
                        # 원본 컬럼명을 한국어 컬럼명으로 매핑
                        row_data = {column_mapping.get(k, k): r.get(k) for k in cols}
                        # integration_type 값 한국어 변환 (integration_type_mapping 사용)
                        if 'integration_type' in r and r['integration_type'] in integration_type_mapping:
                            korean_col_name = column_mapping.get('integration_type', 'integration_type')
                            if korean_col_name in row_data:
                                row_data[korean_col_name] = integration_type_mapping[r['integration_type']]
                        elif '연동유형' in row_data and row_data['연동유형'] in integration_type_mapping:
                            row_data['연동유형'] = integration_type_mapping[row_data['연동유형']]
                        writer.writerow(row_data)
                    # UTF-8 BOM
                    zf.writestr(path, '\ufeff' + sio.getvalue())
                except Exception as e_csv:
                    zf.writestr(path + '.error.txt', str(e_csv))

            if fmt in ['csv','both']:
                write_csv('data/sites.csv', data_sites)
                write_csv('data/site_contacts.csv', data_contacts)
                write_csv('data/site_contact_people.csv', data_contact_people)
                write_csv('data/site_products.csv', data_products)
                write_csv('data/work_items.csv', data_work_items)
                write_csv('data/site_photos.csv', data_photos)
                write_csv('data/site_household_integrations.csv', data_household_integrations)
                write_csv('data/site_common_integrations.csv', data_common_integrations)

            # Excel 현장별 별도 파일 생성 (구조화된 양식)
            if fmt in ['xlsx','both']:
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                    
                    # 각 현장별로 별도 엑셀 파일 생성
                    excel_files_created = 0
                    excel_errors = []
                    
                    for site_id in site_ids:
                        try:
                            # 현장 기본 정보 찾기
                            site_info = next((s for s in data_sites if s.get('id') == site_id), None)
                            if not site_info:
                                excel_errors.append(f"현장 {site_id}: 기본 정보를 찾을 수 없습니다.")
                                continue
                            
                            # 현장별 데이터 수집
                            site_contact = next((c for c in data_contacts if c.get('site_id') == site_id), {})
                            site_product = next((p for p in data_products if p.get('site_id') == site_id), {})
                            site_household = [h for h in data_household_integrations if h.get('site_id') == site_id]
                            site_common = [c for c in data_common_integrations if c.get('site_id') == site_id]
                            
                            # 엑셀 워크북 생성
                            wb = openpyxl.Workbook()
                            ws = wb.active
                            ws.title = "현장 관리 시트"
                            
                            # 스타일 및 레이아웃 정의
                            title_font = Font(name='맑은 고딕', size=16, bold=True)
                            section_font = Font(name='맑은 고딕', size=12, bold=True)
                            label_font = Font(name='맑은 고딕', size=10, bold=True)
                            normal_font = Font(name='맑은 고딕', size=10)
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            columns = ['A', 'B', 'C', 'D', 'E', 'F']
                            col_widths = [16, 22, 16, 22, 16, 22]
                            label_fill = PatternFill(start_color='FFE7E6E6', end_color='FFE7E6E6', fill_type='solid')
                            section_fill = PatternFill(start_color='FFD0CECE', end_color='FFD0CECE', fill_type='solid')

                            ws.sheet_view.showGridLines = False
                            for letter, width in zip(columns, col_widths):
                                ws.column_dimensions[letter].width = width

                            def normalize(value):
                                if value is None:
                                    return ''
                                # 숫자나 날짜는 문자열로 변환하지 않고 그대로 유지
                                if isinstance(value, (int, float, bool)):
                                    return value
                                # 문자열이면 공백 제거
                                if isinstance(value, str):
                                    return value.strip()
                                return value

                            def as_yes_no(value):
                                if value is None:
                                    return ''
                                if isinstance(value, bool):
                                    return '예' if value else '아니오'
                                val = str(value).strip().upper()
                                if val in ('Y', 'YES', 'TRUE', '1'):
                                    return '예'
                                if val in ('N', 'NO', 'FALSE', '0'):
                                    return '아니오'
                                return str(value)

                            def write_label_value(row_idx: int, col_idx: int, label: str, value, span: int = 1, wrap: bool = False) -> int:
                                if col_idx >= len(columns):
                                    return len(columns)
                                label_cell = ws[f'{columns[col_idx]}{row_idx}']
                                label_cell.value = label or ''
                                label_cell.font = label_font
                                label_cell.border = thin_border
                                label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                # openpyxl에서는 fill에 None을 할당할 수 없으므로 조건부로 설정
                                if label:
                                    label_cell.fill = label_fill

                                value_start = col_idx + 1
                                available = len(columns) - value_start
                                if available <= 0:
                                    return len(columns)
                                span = max(1, min(span, available))
                                value_end = value_start + span - 1

                                for idx in range(value_start, value_end + 1):
                                    ws[f'{columns[idx]}{row_idx}'].border = thin_border

                                if value_start != value_end:
                                    ws.merge_cells(f'{columns[value_start]}{row_idx}:{columns[value_end]}{row_idx}')

                                value_cell = ws[f'{columns[value_start]}{row_idx}']
                                value_cell.value = normalize(value)
                                value_cell.font = normal_font
                                value_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=wrap)
                                if wrap:
                                    current_height = ws.row_dimensions[row_idx].height or 0
                                    ws.row_dimensions[row_idx].height = max(current_height, 36)

                                return value_end + 1

                            def write_row_with_pairs(row_idx: int, items: list) -> int:
                                col_idx = 0
                                for item in items:
                                    if col_idx >= len(columns):
                                        break
                                    if item is None:
                                        col_idx = write_label_value(row_idx, col_idx, '', '', span=1)
                                        continue
                                    label = item.get('label', '')
                                    value = item.get('value')
                                    span = item.get('span', 1)
                                    wrap = item.get('wrap', False)
                                    col_idx = write_label_value(row_idx, col_idx, label, value, span=span, wrap=wrap)
                                while col_idx < len(columns):
                                    col_idx = write_label_value(row_idx, col_idx, '', '', span=1)
                                return row_idx + 1

                            def write_section_header(row_idx: int, title: str) -> int:
                                ws.merge_cells(f'A{row_idx}:F{row_idx}')
                                cell = ws[f'A{row_idx}']
                                cell.value = title
                                cell.font = section_font
                                cell.fill = section_fill
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                                cell.border = thin_border
                                for col in columns[1:]:
                                    ws[f'{col}{row_idx}'].border = thin_border
                                return row_idx + 1

                            def write_integration_rows(row_idx: int, records: list, order: list) -> int:
                                records_by_type = {}
                                for rec in records:
                                    key = rec.get('integration_type') if rec else None
                                    key = key or '기타'
                                    records_by_type.setdefault(key, []).append(rec)
                                processed = set()
                                for key in order:
                                    processed.add(key)
                                    entries = records_by_type.get(key)
                                    if not entries:
                                        entries = [None]
                                    for rec in entries:
                                        enabled_val = as_yes_no((rec or {}).get('enabled')) or '예 / 아니오'
                                        row_idx = write_row_with_pairs(row_idx, [
                                            {'label': integration_type_mapping.get(key, key), 'value': enabled_val},
                                            {'label': '업체명', 'value': (rec or {}).get('company_name')},
                                            {'label': '담당자', 'value': (rec or {}).get('contact_person')},
                                        ])
                                        row_idx = write_row_with_pairs(row_idx, [
                                            {'label': '연락처', 'value': (rec or {}).get('contact_phone')},
                                            {'label': '기타', 'value': (rec or {}).get('notes'), 'span': 3, 'wrap': True},
                                        ])
                                for key, entries in records_by_type.items():
                                    if key in processed:
                                        continue
                                    for rec in entries:
                                        enabled_val = as_yes_no((rec or {}).get('enabled')) or '예 / 아니오'
                                        label = integration_type_mapping.get(key, key or '기타')
                                        row_idx = write_row_with_pairs(row_idx, [
                                            {'label': label, 'value': enabled_val},
                                            {'label': '업체명', 'value': (rec or {}).get('company_name')},
                                            {'label': '담당자', 'value': (rec or {}).get('contact_person')},
                                        ])
                                        row_idx = write_row_with_pairs(row_idx, [
                                            {'label': '연락처', 'value': (rec or {}).get('contact_phone')},
                                            {'label': '기타', 'value': (rec or {}).get('notes'), 'span': 3, 'wrap': True},
                                        ])
                                return row_idx

                            ws.merge_cells('A1:F1')
                            ws['A1'] = '현장 관리 시트'
                            ws['A1'].font = title_font
                            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                            ws.row_dimensions[1].height = 28

                            row = 2

                            address = site_info.get('address', '')
                            detail_address = site_info.get('detail_address', '')
                            full_address = f"{address} {detail_address}".strip() if detail_address else address
                            special_notes = site_info.get('special_notes', '')

                            row = write_row_with_pairs(row, [
                                {'label': '등록일', 'value': site_info.get('registration_date')},
                                {'label': '납품예정', 'value': site_info.get('delivery_date')},
                                {'label': '준공일', 'value': site_info.get('completion_date')},
                            ])
                            row = write_row_with_pairs(row, [
                                {'label': '프로젝트 No.', 'value': site_info.get('project_no')},
                                {'label': '건설사', 'value': site_info.get('construction_company')},
                                {'label': '현장명', 'value': site_info.get('site_name')},
                            ])
                            row = write_row_with_pairs(row, [
                                {'label': '주소', 'value': full_address, 'span': 5, 'wrap': True},
                            ])
                            row = write_row_with_pairs(row, [
                                {'label': '세대수', 'value': site_info.get('household_count')},
                                {'label': '제품 BI', 'value': site_info.get('product_bi')},
                                {'label': '인증심사여부', 'value': as_yes_no(site_info.get('certification_audit')) or '예 / 아니오'},
                            ])
                            row = write_row_with_pairs(row, [
                                {'label': 'PM 이름', 'value': site_contact.get('pm_name')},
                                {'label': 'PM 전화번호', 'value': site_contact.get('pm_phone')},
                                {'label': '영업담당자', 'value': site_contact.get('sales_manager_name')},
                            ])
                            row = write_row_with_pairs(row, [
                                {'label': '영업담당자 전화', 'value': site_contact.get('sales_manager_phone')},
                                {'label': '설치점', 'value': site_contact.get('installer_name')},
                                {'label': '설치점 전화', 'value': site_contact.get('installer_phone')},
                            ])
                            row = write_row_with_pairs(row, [
                                {'label': '네트워크점', 'value': site_contact.get('network_manager_name')},
                                {'label': '네트워크점 전화', 'value': site_contact.get('network_manager_phone')},
                                {'label': '건설사 담당자', 'value': site_contact.get('construction_manager_name')},
                            ])
                            row = write_row_with_pairs(row, [
                                {'label': '건설사 담당자 전화', 'value': site_contact.get('construction_manager_phone')},
                                {'label': '외부망 연동', 'value': as_yes_no(site_info.get('external_network_enabled')) or '예 / 아니오'},
                                {'label': '가입기간', 'value': site_info.get('external_network_period')},
                            ])
                            row = write_row_with_pairs(row, [
                                {'label': '홈IoT연동여부', 'value': as_yes_no(site_info.get('home_iot')) or '예 / 아니오'},
                                {'label': '현장관리자', 'value': site_info.get('site_manager')},
                                {'label': '수정일시', 'value': site_info.get('updated_at')},
                            ])
                            row = write_row_with_pairs(row, [
                                {'label': '현장 특이사항', 'value': special_notes, 'span': 5, 'wrap': True},
                            ])

                            row = write_section_header(row, '제품 수량')

                            product_definitions = [
                                ('월패드', 'wallpad_model', 'wallpad_qty'),
                                ('도어폰', 'doorphone_model', 'doorphone_qty'),
                                ('로비폰', 'lobbyphone_model', 'lobbyphone_qty'),
                                ('경비실기', 'guardphone_model', 'guardphone_qty'),
                                ('자석감지기', 'magnet_sensor_model', 'magnet_sensor_qty'),
                                ('동체감지기', 'motion_sensor_model', 'motion_sensor_qty'),
                            ]
                            opener_model = site_product.get('opener_model')
                            opener_qty = site_product.get('opener_qty')
                            if opener_model or opener_qty:
                                product_definitions.append(('개폐기', 'opener_model', 'opener_qty'))

                            for idx in range(0, len(product_definitions), 3):
                                chunk = product_definitions[idx:idx + 3]
                                model_row = []
                                qty_row = []
                                for name, model_key, qty_key in chunk:
                                    model_row.append({'label': name, 'value': site_product.get(model_key)})
                                    qty_row.append({'label': '수량', 'value': site_product.get(qty_key)})
                                while len(model_row) < 3:
                                    model_row.append(None)
                                    qty_row.append(None)
                                row = write_row_with_pairs(row, model_row)
                                row = write_row_with_pairs(row, qty_row)

                            household_order = [
                                'lighting_sw', 'standby_power_sw', 'gas_detector', 'heating', 'ventilation',
                                'door_lock', 'air_conditioner', 'real_time_metering', 'environment_sensor',
                                'vpn', 'all_off_switch', 'bathroom_phone', 'kitchen_tv'
                            ]
                            common_order = [
                                'parking_control', 'remote_metering', 'cctv', 'elevator', 'parcel',
                                'ev_charger', 'parking_location', 'onepass', 'rf_card'
                            ]

                            row = write_section_header(row, '세대부 연동')
                            row = write_integration_rows(row, site_household, household_order)

                            row = write_section_header(row, '공용부 연동')
                            row = write_integration_rows(row, site_common, common_order)

                            # 파일명 생성 (프로젝트No_현장명 형식으로 통일)
                            site_name = str(site_info.get('site_name') or '').strip()
                            if not site_name:
                                site_name = f'현장_{site_id}'
                            
                            project_no = str(site_info.get('project_no') or '').strip()
                            if project_no:
                                # 프로젝트No_현장명 형식
                                filename = f"{project_no}_{site_name}"
                            else:
                                # 프로젝트No가 없으면 현장명만 사용
                                filename = site_name
                            
                            # 파일명에 사용할 수 없는 문자 제거 및 정리
                            invalid_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|', '\n', '\r', '\t']
                            for char in invalid_chars:
                                filename = filename.replace(char, '_')
                            # 연속된 언더스코어 제거
                            while '__' in filename:
                                filename = filename.replace('__', '_')
                            # 앞뒤 공백 및 언더스코어 제거
                            filename = filename.strip('_').strip()
                            if not filename:
                                filename = f'현장_{site_id}'
                            
                            # 엑셀 파일을 메모리에 저장
                            excel_buffer = BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)
                            
                            # 파일 크기 확인 (빈 파일 방지)
                            excel_data = excel_buffer.read()
                            if len(excel_data) < 100:  # 최소 Excel 파일 크기 체크
                                raise Exception(f"생성된 Excel 파일이 너무 작습니다 ({len(excel_data)} bytes)")
                            
                            # ZIP에 추가
                            zf.writestr(f'sites/{filename}.xlsx', excel_data)
                            excel_files_created += 1
                            
                        except Exception as site_error:
                            # site_info가 정의되지 않았을 수 있으므로 안전하게 처리
                            site_name_for_error = 'N/A'
                            try:
                                if 'site_info' in locals() and site_info:
                                    site_name_for_error = site_info.get('site_name', 'N/A')
                            except Exception:
                                pass
                            error_msg = f"현장 {site_id} (현장명: {site_name_for_error}) 엑셀 생성 오류: {str(site_error)}"
                            print(error_msg)
                            excel_errors.append(error_msg)
                            import traceback
                            print(traceback.format_exc())
                            continue
                    
                    # Excel 생성 결과 요약 로그 추가
                    if excel_errors:
                        error_log = f"Excel 생성 중 발생한 오류:\n\n"
                        error_log += f"총 {len(site_ids)}개 현장 중 {excel_files_created}개 성공, {len(excel_errors)}개 실패\n\n"
                        error_log += "\n".join(excel_errors)
                        zf.writestr('data/excel_errors.txt', error_log)
                        print(f"[WARN] Excel 생성 중 {len(excel_errors)}개 현장에서 오류 발생")
                    else:
                        print(f"[INFO] 모든 현장({excel_files_created}개)의 Excel 파일이 성공적으로 생성되었습니다.")
                            
                except Exception as e_xlsx:
                    # 실패 시 안내 파일만 기록
                    try:
                        import sys
                        err_text = f"excel_error={e_xlsx}\npython={sys.version}\nexecutable={sys.executable}"
                    except Exception:
                        err_text = str(e_xlsx)
                    zf.writestr('data/export.xlsx.error.txt', err_text)
                    # openpyxl 미설치 등으로 XLSX 생성 실패 시 CSV 대체본 추가
                    try:
                        from io import StringIO
                        sio = StringIO()
                        df_all.to_csv(sio, index=False)
                        zf.writestr('data/export_fallback.csv', '\ufeff' + sio.getvalue())
                    except Exception:
                        pass

            # 사진 ZIP 포함(원본 다운로드)
            if include_photos and data_photos:
                def in_date_range(uploaded_at_iso: str) -> bool:
                    if not (start_date or end_date):
                        return True
                    try:
                        dt = datetime.fromisoformat((uploaded_at_iso or '').replace('Z','+00:00'))
                    except Exception:
                        return True
                    if start_date:
                        try:
                            s = datetime.fromisoformat(start_date + 'T00:00:00+00:00')
                            if dt < s:
                                return False
                        except Exception:
                            pass
                    if end_date:
                        try:
                            e = datetime.fromisoformat(end_date + 'T23:59:59+00:00')
                            if dt > e:
                                return False
                        except Exception:
                            pass
                    return True

                for ph in data_photos:
                    try:
                        if not in_date_range(str(ph.get('uploaded_at') or '')):
                            continue
                        url = ph.get('image_url')
                        if not url:
                            continue
                        r = requests.get(url, timeout=20)
                        if r.status_code != 200:
                            continue
                        site_id = ph.get('site_id')
                        fname = url.split('/')[-1]
                        yymm = 'unknown'
                        try:
                            dt = datetime.fromisoformat((ph.get('uploaded_at') or '').replace('Z','+00:00'))
                            yymm = f"{dt.year}/{str(dt.month).zfill(2)}"
                        except Exception:
                            pass
                        arcname = f"photos/site_{site_id}/{yymm}/{fname}"
                        zf.writestr(arcname, r.content)
                    except Exception:
                        continue

        buf.seek(0)
        return send_file(buf, mimetype='application/zip', as_attachment=True, download_name=f'export_{ts}.zip')
    except Exception as e:
        return jsonify({'error': str(e)}), 500
# =============================
# 현장별 업무관리: Work Items / Alarms
# =============================

@sites_bp.route('/sites/<int:site_id>/work-items', methods=['GET'])
def list_work_items(site_id):
    try:
        # 인증 체크
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401

        # 권한 확인
        site = supabase.table('sites').select('id, created_by, site_name').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403

        status = (request.args.get('status') or '').strip().lower()
        q = supabase.table('work_items').select('*').eq('site_id', site_id)
        if status in ['todo', 'done']:
            q = q.eq('status', status)
        rows = q.order('id', desc=True).execute()
        return jsonify({'items': rows.data or []}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@sites_bp.route('/sites/<int:site_id>/work-items', methods=['POST'])
def upsert_work_items(site_id):
    """배열 업서트: To do/Done 일괄 저장
    입력 스키마: { items: [ {id?, content, alarm_date?, status('todo'|'done'), done_date?} ] }
    규칙:
      - status=done 저장 시 To do 항목은 status만 'done'으로 업데이트(= To do에서 제외)
      - status=todo 저장 시 alarm_confirmed는 기본 false 유지
    """
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401

        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403

        data = request.get_json() or {}
        items = data.get('items', [])
        if not isinstance(items, list):
            return jsonify({'error': 'items 배열이 필요합니다.'}), 400

        def _to_bool(val):
            if isinstance(val, bool):
                return val
            if isinstance(val, (int, float)):
                return val != 0
            if isinstance(val, str):
                return val.strip().lower() in ('true','1','y','yes','on')
            return False

        saved = []
        deleted_ids = []
        for it in items:
            item_id = it.get('id')
            delete_flag = _to_bool(it.get('delete_flag') or it.get('delete') or it.get('remove'))
            if delete_flag:
                if item_id:
                    try:
                        supabase.table('work_items').delete().eq('id', item_id).eq('site_id', site_id).execute()
                        deleted_ids.append(item_id)
                        print(f"🗑️ 작업 항목 삭제 완료 (id={item_id})")
                    except Exception as delete_err:
                        print(f"❌ 작업 항목 삭제 오류 (id={item_id}): {str(delete_err)}")
                        raise
                else:
                    print("⚠️ ID가 없는 항목 삭제 요청 무시")
                continue

            content = (it.get('content') or '').strip()
            if not content:
                continue
            status = (it.get('status') or 'todo').strip().lower()
            if status not in ['todo','done']:
                status = 'todo'
            # alarm_date 처리: 빈 문자열, None, null을 모두 None으로 변환
            alarm_date_raw = it.get('alarm_date')
            alarm_date = None
            if alarm_date_raw is not None and alarm_date_raw != '':
                # None이 아니고 빈 문자열도 아닌 경우에만 처리
                alarm_date_str = str(alarm_date_raw).strip()
                alarm_date = alarm_date_str if alarm_date_str else None
            
            # payload_data 구성
            # alarm_date는 값이 있을 때만 포함 (None인 경우 제외하여 기존 값 유지)
            # 하지만 alarm_date를 NULL로 설정하려면 명시적으로 포함해야 하므로,
            # 업데이트 시에는 항상 포함하고, 업데이트 후 확인
            payload_data = {
                'site_id': site_id,
                'content': content,
                'status': status,
                'done_date': (it.get('done_date') or None),
                'updated_at': datetime.utcnow().isoformat(),
                'created_by': payload['user_id']
            }
            
            # alarm_date는 명시적으로 포함 (None이어도)
            # Supabase는 None 값을 포함하면 해당 필드를 업데이트하지 않을 수 있으므로,
            # 업데이트 후 실제 데이터를 다시 조회하여 확인
            if alarm_date is not None:
                payload_data['alarm_date'] = alarm_date
            else:
                # alarm_date가 None인 경우, 명시적으로 NULL로 설정하기 위해 포함
                # Supabase Python 클라이언트가 None을 필터링할 수 있으므로,
                # 업데이트 후 확인이 필요
                payload_data['alarm_date'] = None
            
            # 디버깅: 업데이트할 데이터 로그 출력
            if it.get('id'):
                print(f"📝 업데이트할 항목 (id={it.get('id')}): alarm_date={alarm_date}, status={status}, content={content[:50]}")
            # done 저장인데 done_date가 없으면 클라이언트 로컬 날짜를 못받은 경우를 대비해 서버 날짜로 보정
            if status == 'done' and not payload_data['done_date']:
                payload_data['done_date'] = date.today().isoformat()

            # todo 상태인 경우 새 알람은 미확인으로 유지
            if status == 'todo':
                payload_data['alarm_confirmed'] = False

            if it.get('id'):
                # 업데이트 (상태 전환 포함)
                # Supabase는 None 값을 포함한 필드를 업데이트하지 않을 수 있으므로,
                # alarm_date가 None인 경우 명시적으로 처리
                update_payload = dict(payload_data)
                # None 값을 포함한 필드도 업데이트되도록 보장
                # Supabase Python 클라이언트는 None 값을 포함하면 해당 필드를 업데이트하지 않으므로,
                # 업데이트 후 실제 데이터를 다시 조회하여 확인
                try:
                    res = supabase.table('work_items').update(update_payload).eq('id', it['id']).eq('site_id', site_id).execute()
                    # 업데이트 후 실제 데이터를 다시 조회하여 확인 (res.data가 비어있을 수 있음)
                    if res.data and len(res.data) > 0:
                        saved.append(res.data[0])
                        # alarm_date가 None으로 설정되어야 하는데 업데이트되지 않은 경우 재시도
                        updated_item = res.data[0]
                        if alarm_date is None and updated_item.get('alarm_date') is not None:
                            print(f"⚠️ alarm_date가 NULL로 설정되지 않음 (res.data 있음), 재시도...")
                            retry_payload = {'alarm_date': None, 'updated_at': datetime.utcnow().isoformat()}
                            try:
                                retry_res = supabase.table('work_items').update(retry_payload).eq('id', it['id']).eq('site_id', site_id).execute()
                                if retry_res.data:
                                    saved[-1] = retry_res.data[0]
                                    print(f"✅ alarm_date NULL 설정 성공")
                            except Exception as retry_err:
                                print(f"⚠️ alarm_date NULL 설정 재시도 실패: {str(retry_err)}")
                    else:
                        # res.data가 비어있어도 업데이트는 성공했을 수 있으므로, 실제 데이터를 다시 조회
                        verify_res = supabase.table('work_items').select('*').eq('id', it['id']).eq('site_id', site_id).execute()
                        if verify_res.data and len(verify_res.data) > 0:
                            saved.append(verify_res.data[0])
                            # 업데이트된 데이터 확인
                            updated_item = verify_res.data[0]
                            print(f"✅ 업데이트 확인 (id={it['id']}): status={updated_item.get('status')}, alarm_date={updated_item.get('alarm_date')}")
                            # alarm_date가 None으로 설정되어야 하는데 업데이트되지 않은 경우 재시도
                            if alarm_date is None and updated_item.get('alarm_date') is not None:
                                print(f"⚠️ alarm_date가 NULL로 설정되지 않음, 재시도...")
                                # None 값을 명시적으로 NULL로 설정하기 위해 다시 시도
                                retry_payload = {'alarm_date': None, 'updated_at': datetime.utcnow().isoformat()}
                                try:
                                    retry_res = supabase.table('work_items').update(retry_payload).eq('id', it['id']).eq('site_id', site_id).execute()
                                    if retry_res.data:
                                        saved[-1] = retry_res.data[0]  # 업데이트된 데이터로 교체
                                        print(f"✅ alarm_date NULL 설정 성공")
                                except Exception as retry_err:
                                    print(f"⚠️ alarm_date NULL 설정 재시도 실패: {str(retry_err)}")
                        else:
                            print(f"⚠️ 업데이트 확인 실패 (id={it['id']}): 데이터를 찾을 수 없음")
                except Exception as update_err:
                    print(f"❌ 업데이트 오류 (id={it['id']}): {str(update_err)}")
                    raise
            else:
                payload_data['created_at'] = datetime.utcnow().isoformat()
                res = supabase.table('work_items').insert(payload_data).execute()
                if res.data:
                    saved.append(res.data[0])

        return jsonify({'message': '작업 항목이 저장되었습니다.', 'items': saved, 'deleted': deleted_ids}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@sites_bp.route('/sites/<int:site_id>/alarms', methods=['GET'])
def list_alarms(site_id):
    """알람 목록: 조건 alarm_date <= today AND alarm_confirmed = false AND status='todo'"""
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401

        # 권한 확인
        site = supabase.table('sites').select('id, created_by, site_name').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403

        # today는 클라이언트 로컬 날짜(YYYY-MM-DD) 전달 가능, 없으면 서버 날짜 사용
        today = (request.args.get('today') or date.today().isoformat())

        rows = supabase.table('work_items').select('*') \
            .eq('site_id', site_id) \
            .eq('status', 'todo') \
            .eq('alarm_confirmed', False) \
            .lte('alarm_date', today) \
            .order('id', desc=True).execute()

        items = rows.data or []
        # site_name 포함
        for it in items:
            it['site_name'] = site_info.get('site_name')
        return jsonify({'items': items, 'count': len(items)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@sites_bp.route('/sites/<int:site_id>/alarms/confirm', methods=['POST'])
def confirm_alarms(site_id):
    """체크된 알람을 확인 처리: 목록에서 제거되지만 원본의 alarm_date는 유지하고 alarm_confirmed=True로 설정"""
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401

        # 권한 확인
        site = supabase.table('sites').select('id, created_by').eq('id', site_id).execute()
        if not site.data:
            return jsonify({'error': '현장을 찾을 수 없습니다.'}), 404
        site_info = site.data[0]
        if payload['user_role'] != 'admin' and site_info['created_by'] != payload['user_id']:
            return jsonify({'error': '접근 권한이 없습니다.'}), 403

        data = request.get_json() or {}
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'message': '확인할 항목이 없습니다.', 'updated': 0}), 200
        # 일괄 업데이트
        res = supabase.table('work_items').update({
            'alarm_confirmed': True,
            'updated_at': datetime.utcnow().isoformat()
        }).in_('id', ids).eq('site_id', site_id).execute()
        updated_count = len(res.data or [])
        return jsonify({'message': '알람이 확인 처리되었습니다.', 'updated': updated_count}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 프로젝트 번호 중복 체크
@sites_bp.route('/check-project-no', methods=['POST'])
def check_project_no():
    try:
        print(f"🔍 프로젝트 번호 중복 체크 요청")
        print(f"🔑 인증 헤더: {request.headers.get('Authorization', '없음')}")
        print(f"📝 Content-Type: {request.headers.get('Content-Type', '없음')}")
        print(f"📝 Raw 데이터: {request.get_data()}")
        
        # JSON 데이터 안전하게 파싱
        try:
            data = request.get_json()
            print(f"📝 파싱된 JSON 데이터: {data}")
        except Exception as json_error:
            print(f"❌ JSON 파싱 오류: {json_error}")
            return jsonify({'error': '잘못된 JSON 형식입니다.'}), 400
        
        # 인증 확인
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            print("❌ 인증 헤더 없음")
            return jsonify({'error': '인증 토큰이 필요합니다.'}), 401
        
        token = auth_header.split(' ')[1] if auth_header.startswith('Bearer ') else auth_header
        payload = verify_token(token)
        
        if not payload:
            return jsonify({'error': '유효하지 않은 토큰입니다.'}), 401
        
        data = request.get_json()
        project_no = data.get('project_no')
        
        if not project_no:
            return jsonify({'error': '프로젝트 번호가 필요합니다.'}), 400
        
        # 프로젝트 번호 형식 검증 (NA/XXXX 또는 NE/XXXX)
        import re
        if not re.match(r'^(NA|NE)/\d{4}$', project_no):
            return jsonify({'error': '프로젝트 번호 형식이 올바르지 않습니다. (예: NA/1234, NE/5678)'}), 400
        
        # 중복 체크
        existing = supabase.table('sites').select('id, site_name').eq('project_no', project_no).execute()
        
        # 더미 데이터인 경우 항상 사용 가능으로 처리
        if not supabase_url or not supabase_key:
            return jsonify({
                'is_duplicate': False,
                'message': f'프로젝트 번호 "{project_no}"를 사용할 수 있습니다.'
            }), 200
        elif existing.data:
            return jsonify({
                'is_duplicate': True,
                'message': f'프로젝트 번호 "{project_no}"가 이미 사용 중입니다.',
                'existing_site': existing.data[0]
            }), 200
        else:
            return jsonify({
                'is_duplicate': False,
                'message': f'프로젝트 번호 "{project_no}"를 사용할 수 있습니다.'
            }), 200
            
    except Exception as e:
        print(f"❌ 프로젝트 번호 중복 체크 오류: {str(e)}")
        print(f"🔍 오류 타입: {type(e).__name__}")
        import traceback
        print(f"📚 스택 트레이스: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


